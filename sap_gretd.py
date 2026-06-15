"""
SAP GRETD -> ETDAT Otomasyonu
-------------------------------
Kurulum : pip install pywin32 openpyxl
Calistir: python sap_gretd.py
Gereksinim: SAP acik + Scripting aktif + Excel KAPALI olmali
"""

import win32com.client
import openpyxl
import time
import sys
import os

# ── Sabitler (Script Recorder'dan) ───────────────────────────────────────────
TAB_SHIPPING  = r"wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\06"
GRETD_FIELD   = (
    r"wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\06"
    r"/ssubSUBSCREEN_BODY:SAPMV45A:4403"
    r"/subSUBSCREEN_TC:SAPMV45A:4921"
    r"/tblSAPMV45ATCTRL_UEIN_VERSAND"
    r"/ctxtRV45A-GRETD[3,0]"
)
TAB_OVERVIEW  = r"wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\02"
ETDAT_BASE    = (
    r"wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\02"
    r"/ssubSUBSCREEN_BODY:SAPMV45A:4401"
    r"/subSUBSCREEN_TC:SAPMV45A:4900"
    r"/tblSAPMV45ATCTRL_U_ERF_AUFTRAG"
)
ETDAT_COL = 11
# ─────────────────────────────────────────────────────────────────────────────


def sap_baglan():
    try:
        sap = win32com.client.GetObject("SAPGUI")
        app = sap.GetScriptingEngine
        return app.Children(0).Children(0)
    except Exception as e:
        print(f"[HATA] SAP baglantisi kurulamadi: {e}")
        sys.exit(1)


def sap_one_getir(session):
    """SAP penceresini one getirir."""
    try:
        session.FindById("wnd[0]").SetFocus()
        session.FindById("wnd[0]").Restore()
    except Exception:
        pass


def popup_kapat(session):
    """Acik popup varsa Enter ile kapatir."""
    try:
        while session.Children.Count > 1:
            session.FindById("wnd[1]").SendVKey(0)
            time.sleep(0.3)
    except Exception:
        pass


def islemi_yap(session, vbeln):
    """
    Bir VBELN icin:
      1. VA02 ac
      2. VBELN gir
      3. Shipping tab -> GRETD oku
      4. Item Overview tab -> dolu satirlara ETDAT yaz
      5. Ekranda birak (kaydetme yok - kullanici kaydeder)
    Dondurulen deger: (basarili:bool, mesaj:str)
    """
    try:
        # VA02 ac
        session.StartTransaction("VA02")
        time.sleep(0.5)
        popup_kapat(session)
        sap_one_getir(session)

        # VBELN gir
        session.FindById("wnd[0]/usr/ctxtVBAK-VBELN").Text = vbeln
        time.sleep(0.4)
        session.FindById("wnd[0]").SendVKey(0)   # Enter
        time.sleep(0.8)
        popup_kapat(session)

        # Ekran kontrolu
        baslik = session.FindById("wnd[0]").Text
        if "Change Sales Order" not in baslik:
            return False, f"Beklenmeyen ekran: {baslik}"

        # Shipping tab -> GRETD oku
        session.FindById(TAB_SHIPPING).Select()
        time.sleep(0.6)
        gretd = session.FindById(GRETD_FIELD).Text

        if not gretd or not gretd.strip():
            return False, "GRETD BOS"

        print(f"    GRETD okundu: {gretd}")

        # Item Overview tab
        session.FindById(TAB_OVERVIEW).Select()
        time.sleep(0.6)

        # Dolu satirlara ETDAT yaz
        yazilan = 0
        for r in range(99):
            try:
                # Item numarasi bossa satir bos - dur
                posnr = session.FindById(
                    f"{ETDAT_BASE}/txtRV45A-POSNR[0,{r}]"
                ).Text
                if not posnr or not posnr.strip():
                    break

                # ETDAT yaz
                session.FindById(
                    f"{ETDAT_BASE}/ctxtRV45A-ETDAT[{ETDAT_COL},{r}]"
                ).Text = gretd
                yazilan += 1
                print(f"    Satir {r+1} (POSNR={posnr.strip()}): ETDAT = {gretd}")

            except Exception:
                # Satir yok ya da alan yok - dur
                break

        if yazilan == 0:
            return False, "ETDAT YAZILMADI (satir bulunamadi)"

        # *** KAYDETME YOK — kullanici SAP'de Ctrl+S ile kaydedecek ***
        return True, f"YAZILDI — {yazilan} satir, GRETD={gretd} — sen kaydet"

    except Exception as e:
        return False, f"HATA: {e}"


def main():
    # Excel yolu
    if len(sys.argv) > 1:
        excel_yolu = sys.argv[1]
    else:
        excel_yolu = input("Excel dosyasinin tam yolunu girin: ").strip().strip('"')

    if not os.path.exists(excel_yolu):
        print(f"[HATA] Dosya bulunamadi: {excel_yolu}")
        sys.exit(1)

    # SAP baglan
    session = sap_baglan()
    print("[OK] SAP baglantisi kuruldu")

    # Excel ac
    wb = openpyxl.load_workbook(excel_yolu)
    ws = wb.active
    print(f"[OK] Excel acildi: {ws.max_row - 1} satir bulundu")
    print()

    islenen  = 0
    atlanan  = 0
    toplam   = sum(1 for r in ws.iter_rows(min_row=2) if r[0].value)

    for idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
        vbeln  = str(row[0].value).strip() if row[0].value else ""
        durum  = str(row[1].value).strip().lower() if row[1].value else ""

        if not vbeln or vbeln == "None":
            continue

        if durum == "ok":
            print(f"[{idx}/{toplam}] ATLANDI (zaten ok): {vbeln}")
            atlanan += 1
            continue

        print(f"[{idx}/{toplam}] Isleniyor: {vbeln} ...")
        basarili, mesaj = islemi_yap(session, vbeln)
        row[1].value = mesaj
        wb.save(excel_yolu)
        islenen += 1

        if basarili:
            print(f"    [✓] {mesaj}")
            print("    SAP'de Ctrl+S ile kaydet.")
            print("    Kaydettikten sonra ENTER'a bas (CTRL+C = tamamen dur)")
            try:
                input("    > ")
            except KeyboardInterrupt:
                print("\n[DURDURULDU]")
                break
        else:
            print(f"    [✗] {mesaj}")
        print()

    print("=" * 50)
    print(f"BITTI — Islenen: {islenen}  Atlanan: {atlanan}")
    print("=" * 50)


if __name__ == "__main__":
    main()
